from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
from django.template.loader import render_to_string
from .models import Venta
from detalles_venta.models import DetalleVenta
from .forms import VentaForm
from detalles_venta.forms import DetalleVentaForm
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import FileResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.db.models import Q
from django.core.paginator import Paginator
from django.forms import inlineformset_factory
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


def lista_ventas(request):
    ventas = Venta.objects.all().order_by('-fecha')
    
    # Búsqueda
    query = request.GET.get('q')
    if query:
        ventas = ventas.filter(
            Q(cliente__icontains=query) |
            Q(producto__nombre__icontains=query) |
            Q(id__icontains=query)
        )
    
    # Filtrado
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    if fecha_desde:
        ventas = ventas.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(fecha__lte=fecha_hasta)
    
    # Paginación
    paginator = Paginator(ventas, 10)  # 10 ventas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'ventas': page_obj,
        'query': query,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'lista_ventas.html', context)

def editar_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == 'POST':
        form = VentaForm(request.POST, instance=venta)
        if form.is_valid():
            form.save()
            return redirect('lista_ventas')
    else:
        form = VentaForm(instance=venta)
    return render(request, 'editar_venta.html', {'form': form, 'venta': venta})
def crear_venta(request):
    if request.method == 'POST':
        form = VentaForm(request.POST)
        if form.is_valid():
            venta = form.save()
            # Redirigir a la página de crear detalle de venta
            return redirect('crear_detalle_venta', venta_id=venta.id)
    else:
        form = VentaForm()
    return render(request, 'crear_venta.html', {'form': form})

def eliminar_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == 'POST':
        venta.delete()
        return redirect('lista_ventas')
    return render(request, 'eliminar_venta.html', {'venta': venta})
def exportar_excel(request):
    ventas = Venta.objects.all()
    
    # Crear libro y hoja de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    # Agregar encabezados
    headers = ['ID', 'Fecha', 'Cliente', 'Producto', 'Cantidad', 'Precio Unitario', 'Total']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Agregar filas de ventas
    for row_num, venta in enumerate(ventas, 2):
        ws[f'A{row_num}'] = venta.id
        ws[f'B{row_num}'] = venta.fecha.strftime('%d/%m/%Y')
        ws[f'C{row_num}'] = venta.cliente
        ws[f'D{row_num}'] = venta.producto.nombre if venta.producto else 'No especificado'
        ws[f'E{row_num}'] = venta.cantidad
        ws[f'F{row_num}'] = venta.precio_unitario
        ws[f'G{row_num}'] = venta.total

    # Crear respuesta HTTP para descargar el archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
    wb.save(response)
    return response

def exportar_pdf(request):
    # Crear un buffer de bytes para el PDF
    buffer = BytesIO()

    # Crear el documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Obtener los datos de las ventas
    ventas = Venta.objects.all()

    # Crear los datos para la tabla
    data = [['ID', 'Cliente', 'Fecha', 'Total']]  # Encabezados
    for venta in ventas:
        data.append([str(venta.id), venta.cliente, str(venta.fecha), str(venta.total)])

    # Crear la tabla
    table = Table(data)
    
    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # Agregar la tabla al documento
    elements.append(table)

    # Construir el PDF
    doc.build(elements)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='ventas.pdf')





